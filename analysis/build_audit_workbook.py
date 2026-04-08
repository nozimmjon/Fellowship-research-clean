import csv
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


BASE_DIR = os.path.join("outputs", "tables", "audit_two_pass")
OUTPUT_XLSX = os.path.join(BASE_DIR, "two_pass_audit_sheet.xlsx")

SHEETS = [
    ("Replication", "tab1_replication.csv"),
    ("Data_Processed", "tab2_data_processed.csv"),
    ("Outputs", "tab3_outputs.csv"),
    ("Paper_Claims", "tab4_paper_claims.csv"),
    ("Pass1_Summary", "pass1_summary.csv"),
    ("Pass2_Methodology", "pass2_methodology.csv"),
]


def autosize_columns(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 80)


def write_csv_to_sheet(ws, csv_path):
    with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        for row_idx, row in enumerate(reader, start=1):
            ws.append(row)
            if row_idx == 1:
                for cell in ws[row_idx]:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                for cell in ws[row_idx]:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    autosize_columns(ws)


def main():
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    for sheet_name, csv_name in SHEETS:
        ws = wb.create_sheet(title=sheet_name)
        csv_path = os.path.join(BASE_DIR, csv_name)
        if os.path.exists(csv_path):
            write_csv_to_sheet(ws, csv_path)
        else:
            ws.append(["note"])
            ws.append([f"Missing source file: {csv_path}"])

    wb.save(OUTPUT_XLSX)
    print(OUTPUT_XLSX)


if __name__ == "__main__":
    main()
